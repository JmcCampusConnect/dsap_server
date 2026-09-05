from rest_framework import viewsets, status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser

from django.db.models import Q
from django.db import transaction
from django.http import HttpResponse

import openpyxl

from apps.departments.models import ServiceDepartment
from apps.departments.serializers import ServiceDepartmentSerializer
from apps.accounts.models import User
from apps.accounts.permissions import IsSystemAdmin, IsOwnServiceDepartment
from apps.accounts.role_constants import Roles
from apps.audit.models import AuditLog
from common.pagination import StandardPagination


class ServiceDepartmentViewSet(viewsets.ModelViewSet):
    serializer_class = ServiceDepartmentSerializer
    pagination_class = StandardPagination

    def get_queryset(self):
        user = self.request.user
        
        # Start with base queryset with search/filter capabilities
        qs = ServiceDepartment.objects.all().order_by("code")
        
        # Apply search/filter parameters (only for system admin or full access)
        # For service department roles, we'll filter by their department after applying search
        search = self.request.query_params.get("search", "").strip()
        if search:
            qs = qs.filter(
                Q(code__icontains=search)
                | Q(name__icontains=search)
                | Q(status__icontains=search)
                | Q(hod_user_id__username__icontains=search)
                | Q(hod_user_id__email__icontains=search)
            )

        code = self.request.query_params.get("code", "").strip()
        if code:
            qs = qs.filter(code=code)

        name = self.request.query_params.get("name", "").strip()
        if name:
            qs = qs.filter(name=name)

        status_param = self.request.query_params.get("status", "").strip().lower()
        if status_param:
            qs = qs.filter(status__iexact=status_param)

        hod_user_id = self.request.query_params.get("hod_user_id", "").strip()
        if hod_user_id:
            qs = qs.filter(hod_user_id=hod_user_id)

        # Apply role-based filtering
        if user.is_system_admin():
            return qs
        # For service department roles, show only their own department
        if user.has_any_role([Roles.SERVICE_DEPT_ADMIN, Roles.SERVICE_DEPT_STAFF]):
            dept_id = getattr(user.service_department_id, 'id', None)
            if dept_id:
                return qs.filter(id=dept_id)
            return qs.none()
        # Others see nothing (but they won't pass permission anyway)
        return qs.none()
    
    def get_permissions(self):
        # List and retrieve: require authenticated and permission to view own dept(or system admin)
        if self.action in ['list', 'retrieve']:
            return [IsAuthenticated(), IsOwnServiceDepartment()]
        # Create, destroy: only System Admin
        if self.action in ['create', 'destroy']:
            return [IsAuthenticated(), IsSystemAdmin()]
        # Update, partial_update: System Admin (full) or Dept Admin (restricted via serializer)
        return [IsAuthenticated(), IsOwnServiceDepartment()]

    def perform_create(self, serializer):
        instance = serializer.save()
        AuditLog.log(
            request=self.request,
            action="CREATE",
            obj=instance,
            changes=self.get_serializer(instance).data,
        )

    def perform_update(self, serializer):
        instance = self.get_object()
        old_data = self.get_serializer(instance).data
        updated_instance = serializer.save()
        new_data = self.get_serializer(updated_instance).data

        changes = {}
        status_changed = False
        old_status = old_data.get('status', '').lower()
        new_status = new_data.get('status', '').lower()
        
        for key, new_value in new_data.items():
            old_value = old_data.get(key)
            if old_value != new_value:
                changes[key] = {"old": old_value, "new": new_value}
                if key == 'status':
                    status_changed = True

        # Log as ACTIVATE or DEACTIVATE if only status is changing
        if status_changed and len(changes) == 1:
            if old_status == 'active' and new_status == 'inactive':
                action = "DEACTIVATE"
            elif old_status == 'inactive' and new_status == 'active':
                action = "ACTIVATE"
            else:
                action = "UPDATE"
        else:
            action = "UPDATE"

        AuditLog.log(
            request=self.request,
            action=action,
            obj=updated_instance,
            changes=changes,
        )

    def destroy(self, request, *args, **kwargs):
        """
        Soft delete (Deactivate) - only System Admin can call this (permission enforced).
        """
        department = self.get_object()
        snapshot = self.get_serializer(department).data
        object_id = department.pk
        department.status = "inactive"
        department.save()

        AuditLog.log(
            request=request,
            action="DEACTIVATE",
            obj=department,
            object_id=object_id,
            changes=snapshot,
        )

        return Response(
            {"message": "Service Department deactivated successfully."},
            status=status.HTTP_200_OK
        )

    @action(detail=False, methods=["get"], url_path="options")
    def get_options(self, request):
        """
        Dropdown options for the service department filter UI.
        """
        # Apply permission filtering for options
        user = request.user
        base_qs = ServiceDepartment.objects.all()
        
        if not user.is_system_admin():
            if user.has_any_role([Roles.SERVICE_DEPT_ADMIN, Roles.SERVICE_DEPT_STAFF]):
                dept_id = getattr(user.service_department_id, 'id', None)
                if dept_id:
                    base_qs = base_qs.filter(id=dept_id)
                else:
                    base_qs = base_qs.none()

        codes = list(
            base_qs.exclude(code="").values_list("code", flat=True).distinct().order_by("code")
        )
        names = list(
            base_qs.exclude(name="").values_list("name", flat=True).distinct().order_by("name")
        )

        # HOD options: id + display label from related user
        hod_entries = (
            base_qs.exclude(hod_user_id=None)
            .select_related("hod_user_id")
            .values(
                "hod_user_id__id",
                "hod_user_id__username",
                "hod_user_id__email",
            )
            .distinct()
            .order_by("hod_user_id__username")
        )
        hod_options = [
            {
                "value": str(entry["hod_user_id__id"]),
                "label": entry["hod_user_id__username"] or entry["hod_user_id__email"] or str(entry["hod_user_id__id"]),
            }
            for entry in hod_entries
        ]

        return Response({
            "codes": [{"value": c, "label": c} for c in codes],
            "names": [{"value": n, "label": n} for n in names],
            "hods": hod_options,
            "statuses": [
                {"value": "active", "label": "Active"},
                {"value": "inactive", "label": "Inactive"},
            ],
        })

    @action(detail=False, methods=["get"], url_path="export")
    def export_excel(self, request):
        qs = self.get_queryset().select_related("hod_user_id")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Service Departments"

        headers = ["Code", "Name", "HOD Username", "Status"]
        ws.append(headers)

        for department in qs:
            hod_username = department.hod_user_id.username if department.hod_user_id else ""
            ws.append([
                department.code,
                department.name,
                hod_username,
                department.status.lower(),
            ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=service_departments.xlsx"
        wb.save(response)

        AuditLog.log(
            request=request,
            action="EXPORT",
            obj=ServiceDepartment,
            object_id="EXPORT_SERVICE_DEPARTMENTS",
            object_repr="Exported Service Departments",
            changes={"exported_count": qs.count()},
        )

        return response

    @action(detail=False, methods=["post"], url_path="import", parser_classes=[MultiPartParser, FormParser])
    def import_excel(self, request):
        # Only system admin can import
        if not request.user.is_system_admin():
            return Response(
                {"error": "Only system administrators can import departments."},
                status=status.HTTP_403_FORBIDDEN
            )

        file = request.FILES.get("file")
        if not file:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
        except Exception:
            return Response({"error": "Invalid Excel file format"}, status=status.HTTP_400_BAD_REQUEST)

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return Response({"error": "File is empty or contains only headers"}, status=status.HTTP_400_BAD_REQUEST)

        headers = [str(h).strip() for h in rows[0] if h is not None]
        expected_headers = ["Code", "Name", "HOD Username", "Status"]

        if len(headers) < len(expected_headers) or headers[:len(expected_headers)] != expected_headers:
            return Response(
                {"error": f"Invalid headers. Expected: {', '.join(expected_headers)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        errors = []
        valid_departments = []
        db_existing_codes = set(ServiceDepartment.objects.values_list("code", flat=True))
        seen_codes_in_file = set()

        for row_idx, row in enumerate(rows[1:], start=2):
            if not any(row):
                continue

            code = str(row[0]).strip().upper() if row[0] is not None else ""
            name = str(row[1]).strip() if row[1] is not None else ""
            hod_username = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            status_value = str(row[3]).strip().lower() if len(row) > 3 and row[3] is not None else "active"

            row_errors = []
            if not code:
                row_errors.append("Department Code is required.")
            elif len(code) > 20:
                row_errors.append("Department Code cannot exceed 20 characters.")
            elif code in db_existing_codes:
                row_errors.append(f"Department Code '{code}' already exists.")
            elif code in seen_codes_in_file:
                row_errors.append(f"Duplicate Department Code '{code}' found within the uploaded Excel file.")
            else:
                seen_codes_in_file.add(code)

            if not name:
                row_errors.append("Department Name is required.")
            elif len(name) > 150:
                row_errors.append("Department Name exceeds the maximum allowed length.")

            if status_value not in {"active", "inactive"}:
                row_errors.append("Status must be active or inactive.")

            hod_user_obj = None
            if hod_username:
                try:
                    user = User.objects.get(username=hod_username)
                    if user.role_id.name != "SERVICE_DEPT_ADMIN":
                        row_errors.append(f"User '{hod_username}' is not a Service Department Admin.")
                    else:
                        hod_user_obj = user
                except User.DoesNotExist:
                    row_errors.append(f"HOD username '{hod_username}' does not exist.")

            if row_errors:
                errors.append({"row": row_idx, "errors": row_errors})
            else:
                valid_departments.append(
                    ServiceDepartment(
                        code=code,
                        name=name,
                        hod_user_id=hod_user_obj,
                        status=status_value,
                    )
                )

        if errors:
            return Response(
                {"error": "Validation failed for some rows", "details": errors},
                status=status.HTTP_400_BAD_REQUEST,
            )

        with transaction.atomic():
            ServiceDepartment.objects.bulk_create(valid_departments)
            AuditLog.log(
                request=request,
                action="IMPORT",
                obj=ServiceDepartment,
                object_id="BULK_IMPORT",
                object_repr="Imported Service Departments",
                changes={"imported_count": len(valid_departments)},
            )

        return Response(
            {"message": f"Successfully imported {len(valid_departments)} departments"},
            status=status.HTTP_201_CREATED,
        )