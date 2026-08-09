from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny
from django.db.models import Q
from apps.departments.models import AcademicDepartment
from apps.departments.serializers.academic_department import AcademicDepartmentSerializer
from apps.accounts.models import User
from common.pagination import StandardPagination
import openpyxl
from django.http import HttpResponse
from rest_framework.parsers import MultiPartParser, FormParser
from django.db import transaction
from apps.audit.models import AuditLog


class AcademicDepartmentViewSet(viewsets.ModelViewSet):

    serializer_class = AcademicDepartmentSerializer
    pagination_class = StandardPagination

    def get_permissions(self):
        return [AllowAny()]

    def get_queryset(self):
        qs = AcademicDepartment.objects.exclude(status="INACTIVE").order_by("code")

        search = self.request.query_params.get("search", "").strip()
        if search:
            qs = qs.filter(
                Q(code__icontains=search)
                | Q(stream__icontains=search)
                | Q(degree__icontains=search)
                | Q(branch__icontains=search)
                | Q(type__icontains=search)
                | Q(category__icontains=search)
            )

        for field in ("code", "stream", "degree", "branch", "type", "category"):
            val = self.request.query_params.get(field, "").strip()
            if val:
                qs = qs.filter(**{field: val})

        return qs

    def perform_create(self, serializer):
        instance = serializer.save()
        changes = self.get_serializer(instance).data
        AuditLog.log(
            request=self.request,
            action='CREATE',
            obj=instance,
            changes=changes
        )

    def perform_update(self, serializer):
        instance = self.get_object()
        old_data = self.get_serializer(instance).data
        updated_instance = serializer.save()
        new_data = self.get_serializer(updated_instance).data

        changes = {}
        for key, new_value in new_data.items():
            old_value = old_data.get(key)
            if old_value != new_value:
                changes[key] = {'old': old_value, 'new': new_value}

        if changes:
            AuditLog.log(
                request=self.request,
                action='UPDATE',
                obj=updated_instance,
                changes=changes
            )

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        snapshot = self.get_serializer(instance).data
        object_id = instance.pk
        instance.delete()
        AuditLog.log(
            request=request,
            action='DELETE',
            obj=instance,
            object_id=object_id,
            changes=snapshot
        )
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=["get"], url_path="options")
    def get_options(self, request):
        base_qs = AcademicDepartment.objects.exclude(status="INACTIVE")

        stream_filter = request.query_params.get("stream", "").strip()
        type_filter = request.query_params.get("type", "").strip()
        category_filter = request.query_params.get("category", "").strip()
        degree_filter = request.query_params.get("degree", "").strip()
        branch_filter = request.query_params.get("branch", "").strip()

        # All active records for frontend client-side dynamic filtering
        records = list(
            base_qs.values("stream", "type", "category", "degree", "branch", "code").distinct()
        )

        streams = list(base_qs.exclude(stream="").values_list("stream", flat=True).distinct())

        type_qs = base_qs.exclude(type="")
        if stream_filter:
            type_qs = type_qs.filter(stream=stream_filter)
        types = list(type_qs.values_list("type", flat=True).distinct())

        cat_qs = base_qs.exclude(category="")
        if stream_filter:
            cat_qs = cat_qs.filter(stream=stream_filter)
        if type_filter:
            cat_qs = cat_qs.filter(type=type_filter)
        categories = list(cat_qs.values_list("category", flat=True).distinct())

        deg_qs = base_qs.exclude(degree="")
        if stream_filter:
            deg_qs = deg_qs.filter(stream=stream_filter)
        if type_filter:
            deg_qs = deg_qs.filter(type=type_filter)
        if category_filter:
            deg_qs = deg_qs.filter(category=category_filter)
        degrees = list(deg_qs.values_list("degree", flat=True).distinct())

        branch_qs = base_qs.exclude(branch="")
        if stream_filter:
            branch_qs = branch_qs.filter(stream=stream_filter)
        if type_filter:
            branch_qs = branch_qs.filter(type=type_filter)
        if category_filter:
            branch_qs = branch_qs.filter(category=category_filter)
        if degree_filter:
            branch_qs = branch_qs.filter(degree=degree_filter)
        branches = list(branch_qs.values_list("branch", flat=True).distinct())

        code_qs = base_qs.exclude(code="")
        if stream_filter:
            code_qs = code_qs.filter(stream=stream_filter)
        if type_filter:
            code_qs = code_qs.filter(type=type_filter)
        if category_filter:
            code_qs = code_qs.filter(category=category_filter)
        if degree_filter:
            code_qs = code_qs.filter(degree=degree_filter)
        if branch_filter:
            code_qs = code_qs.filter(branch=branch_filter)
        codes = list(code_qs.values_list("code", flat=True).distinct())

        return Response({
            "streams": [{"value": x, "label": x} for x in sorted(streams)],
            "types": [{"value": x, "label": x} for x in sorted(types)],
            "categories": [{"value": x, "label": x} for x in sorted(categories)],
            "degrees": [{"value": x, "label": x} for x in sorted(degrees)],
            "branches": [{"value": x, "label": x} for x in sorted(branches)],
            "codes": [{"value": x, "label": x} for x in sorted(codes)],
            "records": records,
        })

    @action(detail=False, methods=['get'], url_path='export')
    def export_excel(self, request):
        qs = self.get_queryset()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Academic Departments"
        
        headers = ["Stream", "Type", "Category", "Degree", "Branch", "Code"]
        ws.append(headers)
        
        for dept in qs:
            ws.append([
                dept.stream,
                dept.type,
                dept.category,
                dept.degree,
                dept.branch,
                dept.code,
            ])
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=academic_departments.xlsx'
        wb.save(response)
        
        return response

    def _extract_header_map(self, header_row):
        """Map required fields to column indices based on header names."""
        REQUIRED_FIELDS = {"code", "stream", "degree", "branch", "type", "category"}
        col_map = {}
        if not header_row:
            return col_map, REQUIRED_FIELDS

        for idx, cell in enumerate(header_row):
            if cell is None:
                continue
            name = str(cell).strip().lower()
            if name in REQUIRED_FIELDS and name not in col_map:
                col_map[name] = idx

        missing = REQUIRED_FIELDS - set(col_map.keys())
        return col_map, missing

    @action(detail=False, methods=['post'], url_path='parse_excel', parser_classes=[MultiPartParser, FormParser])
    def parse_excel(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
        except Exception:
            return Response({"error": "Invalid Excel file format"}, status=status.HTTP_400_BAD_REQUEST)

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return Response({"error": "File is empty or contains only headers"}, status=status.HTTP_400_BAD_REQUEST)

        col_map, missing = self._extract_header_map(rows[0])
        if missing:
            missing_readable = ", ".join(sorted([m.capitalize() for m in missing]))
            return Response({
                "error": f"Missing required header(s): {missing_readable}. Required headers are: Code, Stream, Degree, Branch, Type, Category."
            }, status=status.HTTP_400_BAD_REQUEST)

        db_existing = set(
            AcademicDepartment.objects.values_list('code', 'stream')
        )
        db_existing_lower = {(c.strip().upper(), s.strip().lower()) for c, s in db_existing}

        seen_pairs_in_file = set()
        parsed_rows = []

        summary = {
            "new": 0,
            "invalid": 0,
            "duplicate": 0,
            "already_exists": 0,
            "total": 0
        }

        for row_idx, row in enumerate(rows[1:], start=2):
            if not any(row):
                continue

            get_val = lambda field: str(row[col_map[field]]).strip() if col_map[field] < len(row) and row[col_map[field]] is not None else ""

            code = get_val("code").upper()
            stream = get_val("stream")
            degree = get_val("degree")
            branch = get_val("branch")
            dept_type = get_val("type")
            category = get_val("category")

            row_data = {
                "row_number": row_idx,
                "code": code,
                "stream": stream,
                "degree": degree,
                "branch": branch,
                "type": dept_type,
                "category": category,
                "status": "NEW",
                "errors": []
            }

            row_errors = []
            if not code:
                row_errors.append("Department Code is required.")
            elif len(code) > 20:
                row_errors.append("Department Code exceeds 20 characters.")

            if not stream:
                row_errors.append("Stream is required.")

            if not degree:
                row_errors.append("Degree is required.")
            elif len(degree) > 50:
                row_errors.append("Degree exceeds 50 characters.")

            if not branch:
                row_errors.append("Branch is required.")
            elif len(branch) > 100:
                row_errors.append("Branch exceeds 100 characters.")

            if not dept_type:
                row_errors.append("Type is required.")
            elif len(dept_type) > 100:
                row_errors.append("Type exceeds 100 characters.")

            if not category:
                row_errors.append("Category is required.")
            elif len(category) > 100:
                row_errors.append("Category exceeds 100 characters.")

            summary["total"] += 1

            if row_errors:
                row_data["status"] = "INVALID"
                row_data["errors"] = row_errors
                summary["invalid"] += 1
            else:
                pair = (code, stream.lower())
                if (code, stream.lower()) in db_existing_lower:
                    row_data["status"] = "ALREADY EXISTS"
                    row_data["errors"] = [f"Department Code '{code}' with Stream '{stream}' already exists in database."]
                    summary["already_exists"] += 1
                elif pair in seen_pairs_in_file:
                    row_data["status"] = "DUPLICATE"
                    row_data["errors"] = [f"Duplicate Department Code '{code}' with Stream '{stream}' in uploaded file."]
                    summary["duplicate"] += 1
                else:
                    seen_pairs_in_file.add(pair)
                    row_data["status"] = "NEW"
                    summary["new"] += 1

            parsed_rows.append(row_data)

        return Response({
            "summary": summary,
            "rows": parsed_rows
        })

    @action(detail=False, methods=['post'], url_path='confirm_import')
    def confirm_import(self, request):
        items = request.data.get('items', [])
        if not items or not isinstance(items, list):
            return Response({"error": "No items provided for import"}, status=status.HTTP_400_BAD_REQUEST)

        db_existing = set(
            (c.strip().upper(), s.strip().lower())
            for c, s in AcademicDepartment.objects.values_list('code', 'stream')
        )

        valid_departments = []
        errors = []

        for idx, item in enumerate(items, start=1):
            serializer = AcademicDepartmentSerializer(data=item)
            if not serializer.is_valid():
                errors.append({"item": idx, "code": item.get('code'), "errors": serializer.errors})
                continue

            code = serializer.validated_data['code']
            stream = serializer.validated_data['stream']
            pair = (code.upper(), stream.lower())

            if pair in db_existing:
                errors.append({"item": idx, "code": code, "errors": f"Department Code '{code}' with Stream '{stream}' already exists."})
                continue

            db_existing.add(pair)
            valid_departments.append(AcademicDepartment(
                code=code,
                stream=stream,
                degree=serializer.validated_data['degree'],
                branch=serializer.validated_data['branch'],
                type=serializer.validated_data['type'],
                category=serializer.validated_data['category'],
            ))

        if errors and not valid_departments:
            return Response({"error": "Validation failed for all items", "details": errors}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            created_objs = AcademicDepartment.objects.bulk_create(valid_departments)
            AuditLog.log(
                request=request,
                action='IMPORT',
                obj=AcademicDepartment,
                object_id='BULK_IMPORT',
                object_repr='Imported Academic Departments',
                changes={"imported_count": len(created_objs)}
            )

        return Response({
            "message": f"Successfully imported {len(created_objs)} departments",
            "imported_count": len(created_objs)
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'], url_path='import', parser_classes=[MultiPartParser, FormParser])
    def import_excel(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
        except Exception:
            return Response({"error": "Invalid Excel file format"}, status=status.HTTP_400_BAD_REQUEST)

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return Response({"error": "File is empty or contains only headers"}, status=status.HTTP_400_BAD_REQUEST)

        col_map, missing = self._extract_header_map(rows[0])
        if missing:
            missing_readable = ", ".join(sorted([m.capitalize() for m in missing]))
            return Response({
                "error": f"Missing required header(s): {missing_readable}. Required headers are: Code, Stream, Degree, Branch, Type, Category."
            }, status=status.HTTP_400_BAD_REQUEST)

        errors = []
        valid_departments = []
        db_existing_pairs = set((c.upper(), s.lower()) for c, s in AcademicDepartment.objects.values_list('code', 'stream'))
        seen_pairs_in_file = set()

        for row_idx, row in enumerate(rows[1:], start=2):
            if not any(row):
                continue

            get_val = lambda field: str(row[col_map[field]]).strip() if col_map[field] < len(row) and row[col_map[field]] is not None else ""

            code = get_val("code").upper()
            stream = get_val("stream")
            degree = get_val("degree")
            branch = get_val("branch")
            dept_type = get_val("type")
            category = get_val("category")

            row_errors = []
            if not stream:
                row_errors.append("Stream is required.")

            pair = (code, stream.lower())
            if not code:
                row_errors.append("Department Code is required.")
            elif len(code) > 20:
                row_errors.append("Department Code cannot exceed 20 characters.")
            elif pair in db_existing_pairs:
                row_errors.append(f"Department Code '{code}' with Stream '{stream}' already exists.")
            elif pair in seen_pairs_in_file:
                row_errors.append(f"Duplicate Department Code '{code}' with Stream '{stream}' found within the uploaded Excel file.")
            else:
                seen_pairs_in_file.add(pair)

            if not degree:
                row_errors.append("Degree is required.")
            elif len(degree) > 50:
                row_errors.append("Degree exceeds the maximum allowed length.")

            if not branch:
                row_errors.append("Branch is required.")
            elif len(branch) > 100:
                row_errors.append("Branch exceeds the maximum allowed length.")

            if not dept_type:
                row_errors.append("Type is required.")
            elif len(dept_type) > 100:
                row_errors.append("Type exceeds the maximum allowed length.")

            if not category:
                row_errors.append("Category is required.")
            elif len(category) > 100:
                row_errors.append("Category exceeds the maximum allowed length.")

            if row_errors:
                errors.append({"row": row_idx, "errors": row_errors})
            else:
                valid_departments.append(AcademicDepartment(
                    code=code,
                    stream=stream,
                    degree=degree,
                    branch=branch,
                    type=dept_type,
                    category=category,
                ))

        if errors:
            return Response({
                "error": "Validation failed for some rows",
                "details": errors
            }, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            created_objs = AcademicDepartment.objects.bulk_create(valid_departments)
            AuditLog.log(
                request=request,
                action='IMPORT',
                obj=AcademicDepartment,
                object_id='BULK_IMPORT',
                object_repr='Imported Academic Departments',
                changes={"imported_count": len(created_objs)}
            )

        return Response({
            "message": f"Successfully imported {len(created_objs)} departments"
        }, status=status.HTTP_201_CREATED)