from django.contrib.auth.hashers import make_password
from django.http import HttpResponse
from django.db import transaction
import openpyxl
from rest_framework import status, viewsets, filters
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from ..models import User
from ..serializers import (
    UserSerializer,
    ResetPasswordSerializer,
)
from apps.audit.models import AuditLog
from ..permissions import IsSystemAdmin, IsServiceDeptAdmin, IsSelfOrSystemAdmin
from ..role_constants import Roles



class UserViewSet(viewsets.ModelViewSet):
    serializer_class = UserSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ["username","email"]

    def get_permissions(self):
        if self.action in ['me','retrieve']:
            return [IsAuthenticated()]
        if self.action in ['reset_password','activate']:
            return [IsAuthenticated(), IsSelfOrSystemAdmin()]
        return [IsAuthenticated(), (IsSystemAdmin | IsServiceDeptAdmin)()]


    def get_queryset(self):
        user = self.request.user
        qs = User.objects.select_related('role_id','service_department_id').all().order_by("id")
        if user.is_system_admin():
            role_id = self.request.query_params.get("role_id")
            is_active = self.request.query_params.get("is_active")
            if role_id:
                qs = qs.filter(role_id_id=role_id)
            if is_active is not None:
                qs = qs.filter(is_active=is_active.lower() == "true")
            return qs
        if user.has_role(Roles.SERVICE_DEPT_ADMIN):
            # Only own department + self
            dept_id = getattr(user.service_department_id, 'id', None) if user.service_department_id else None
            if dept_id:
                return qs.filter(service_department_id_id=dept_id)
            return qs.filter(id=user.id)
        # Staff/Student/Teaching can only see self
        return qs.filter(id=user.id)
    
    def get_object(self):
        obj = super().get_object()
        self.check_object_permissions(self.request, obj)
        return obj

    def perform_create(self, serializer):
        instance = serializer.save()

        changes = self.get_serializer(instance).data

        AuditLog.log(
            request=self.request,
            action="CREATE",
            obj=instance,
            changes=changes,
        )

    def perform_update(self, serializer):
        instance = self.get_object()
       

        old_data = self.get_serializer(instance).data

        updated_instance = serializer.save()

        new_data = self.get_serializer(updated_instance).data

        changes = {}

        for key, new_value in new_data.items():
            old_value = old_data.get(key)

            if old_value != new_value:
                changes[key] = {
                    "old": old_value,
                    "new": new_value,
                }

        if changes:
            
            AuditLog.log(
                request=self.request,
                action="UPDATE",
                obj=updated_instance,
                changes=changes,
            )

    @action(detail=False, methods=["get"], url_path="export")
    def export_excel(self, request):
        qs = self.get_queryset().select_related(
            "role_id",
            "service_department_id",
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"

        headers = [
            "Username",
            "Email",
            "Role",
            "Service Department",
            "Status",
        ]
        ws.append(headers)

        for user in qs:
            role_name = user.role_id.name if user.role_id else ""
            department_name = (
                user.service_department_id.name
                if user.service_department_id
                else ""
            )

            ws.append([
                user.username,
                user.email,
                role_name,
                department_name,
                "Active" if user.is_active else "Inactive",
            ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=users.xlsx"

        wb.save(response)

        AuditLog.log(
            request=request,
            action="EXPORT",
            obj=User,
            object_id="EXPORT_USERS",
            object_repr="Exported Users",
            changes={"exported_count": qs.count()},
        )

        return response
    
    @action(
        detail=False,
        methods=["post"],
        url_path="import",
        parser_classes=[MultiPartParser, FormParser],
    )
    def import_excel(self, request):
        # Only system admin can import users
        if not request.user.is_system_admin():
            return Response(
                {"error": "Only system administrators can import users."},
                status=status.HTTP_403_FORBIDDEN,
            )

        file = request.FILES.get("file")

        if not file:
            return Response(
                {"error": "No file uploaded"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
        except Exception:
            return Response(
                {"error": "Invalid Excel file format"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            return Response(
                {"error": "File is empty or contains only headers"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        headers = [str(h).strip() for h in rows[0] if h is not None]

        expected_headers = [
            "Username",
            "Email",
            "Role",
            "Status",
        ]

        if (
            len(headers) < len(expected_headers)
            or headers[:len(expected_headers)] != expected_headers
        ):
            return Response(
                {
                    "error": (
                        f"Invalid headers. Expected: "
                        f"{', '.join(expected_headers)}"
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        errors = []
        valid_users = []

        db_existing_usernames = set(
            User.objects.values_list("username", flat=True)
        )
        db_existing_emails = set(
            User.objects.exclude(email__isnull=True)
            .exclude(email="")
            .values_list("email", flat=True)
        )

        seen_usernames = set()
        seen_emails = set()

        roles = {
            role.name: role
            for role in User._meta.get_field("role_id").remote_field.model.objects.all()
        }

       

        for row_idx, row in enumerate(rows[1:], start=2):

            if not any(row):
                continue

            username = (
                str(row[0]).strip()
                if row[0] is not None
                else ""
            )

            email = (
                str(row[1]).strip()
                if len(row) > 1 and row[1] is not None
                else ""
            )

            role_name = (
                str(row[2]).strip()
                if len(row) > 2 and row[2] is not None
                else ""
            )

            department_name = (
                str(row[3]).strip()
                if len(row) > 3 and row[3] is not None
                else ""
            )

            status_value = (
                str(row[3]).strip().lower()
                if len(row) > 3 and row[3] is not None
                else "active"
            )

            row_errors = []

            # Username validation
            if not username:
                row_errors.append("Username is required.")
            elif len(username) > 100:
                row_errors.append(
                    "Username cannot exceed 100 characters."
                )
            elif username in db_existing_usernames:
                row_errors.append(
                    f"Username '{username}' already exists."
                )
            elif username in seen_usernames:
                row_errors.append(
                    f"Duplicate Username '{username}' found within the uploaded Excel file."
                )
            else:
                seen_usernames.add(username)

            # Email validation
            if not email:
                row_errors.append("Email is required.")
            elif len(email) > 150:
                row_errors.append(
                    "Email cannot exceed 150 characters."
                )
            elif email in db_existing_emails:
                row_errors.append(
                    f"Email '{email}' already exists."
                )
            elif email in seen_emails:
                row_errors.append(
                    f"Duplicate Email '{email}' found within the uploaded Excel file."
                )
            else:
                seen_emails.add(email)

            # Role validation
            role_obj = None

            if not role_name:
                row_errors.append("Role is required.")
            elif role_name not in roles:
                row_errors.append(
                    f"Role '{role_name}' does not exist."
                )
            else:
                role_obj = roles[role_name]


            # Status validation
            if status_value not in {"active", "inactive"}:
                row_errors.append(
                    "Status must be active or inactive."
                )

            if row_errors:
                errors.append(
                    {
                        "row": row_idx,
                        "errors": row_errors,
                    }
                )
            else:
                valid_users.append(
                    User(
                        username=username,
                        email=email,
                        password_hash="",
                        role_id=role_obj,
                        is_active=status_value == "active",
                    )
                )

        if errors:
            return Response(
                {
                    "error": "Validation failed for some rows",
                    "details": errors,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        with transaction.atomic():

            User.objects.bulk_create(valid_users)

            AuditLog.log(
                request=request,
                action="IMPORT",
                obj=User,
                object_id="BULK_IMPORT",
                object_repr="Imported Users",
                changes={
                    "imported_count": len(valid_users)
                },
            )

        return Response(
            {
                "message": (
                    f"Successfully imported "
                    f"{len(valid_users)} users"
                )
            },
            status=status.HTTP_201_CREATED,
        )

    def perform_destroy(self, instance):
        old_status = instance.is_active

        if instance.id == self.request.user.id:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Cannot deactivate self")
        instance.is_active = False
        instance.save()

        AuditLog.log(
            
            request=self.request,
            action="DEACTIVATE",
            obj=instance,
            changes={
                "is_active": {
                    "old": old_status,
                    "new": False,
                }
            },
        )

    @action(detail=False, methods=["get"], url_path="me")
    def me(self, request):
        serializer = self.get_serializer(request.user)
        return Response(serializer.data)

    @action(detail=True, methods=["post"], url_path="reset-password")
    def reset_password(self, request, pk=None):
        user = self.get_object()

        serializer = ResetPasswordSerializer(
            data=request.data
        )
        serializer.is_valid(raise_exception=True)

        user.password_hash = make_password(
            serializer.validated_data["password"]
        )
        user.save()

        AuditLog.log(
            request=request,
            action="UPDATE",
            obj=user,
            changes={
                "password": "Password reset"
            },
        )

        return Response(
            {"message": "Password reset successfully."},
            status=status.HTTP_200_OK,
        )

    @action(detail=True, methods=["post"], url_path="activate")
    def activate(self, request, pk=None):
        user = self.get_object()

        old_status = user.is_active

        user.is_active = True
        user.save()

        AuditLog.log(
            request=request,
            action="ACTIVATE",
            obj=user,
            changes={
                "is_active": {
                    "old": old_status,
                    "new": True,
                }
            },
        )

        return Response(
            {"message": "User activated successfully."},
            status=status.HTTP_200_OK,
        )
