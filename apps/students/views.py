import csv
import datetime as dt
import io

import openpyxl
from django.contrib.auth.hashers import make_password
from django.core.exceptions import ValidationError as DjangoValidationError
from django.core.validators import validate_email
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from rest_framework import filters, status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.accounts.models import Role, User
from apps.audit.models import AuditLog
from apps.departments.models import AcademicDepartment
from common.pagination import StandardPagination

from .models import Student
from .permissions import IsStudentManagementAdmin
from .serializers import StudentSerializer


class StudentViewSet(viewsets.ModelViewSet):
    """
    Read-only viewset for listing students with search, filters, and pagination.
    """

    serializer_class = StudentSerializer
    pagination_class = StandardPagination
    permission_classes = [IsStudentManagementAdmin]

    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = [
        'register_number',
        'mobile_number',
        'name',
        'year_of_admission',
        'section',
        'stream',
        'academic_department_id__code',
        'academic_department_id__degree',
        'academic_department_id__branch',
    ]
    ordering_fields = ['register_number', 'year_of_admission', 'created_at']
    ordering = ['register_number']

    EXCEL_HEADERS = [
        'Register Number',
        'Name',
        'Email',
        'Year of Admission',
        'DOB (DD-MM-YYYY)',
        'Section',
        'Stream',
        'Mobile Number',
        'Status',
    ]

    HEADER_ALIASES = {
        'register number': 'register_number',
        'name': 'name',
        'email': 'email',
        'year of admission': 'year_of_admission',
        'dob': 'dob',
        'dob dd mm yyyy': 'dob',
        'date of birth': 'dob',
        'dob yyyy mm dd': 'dob',
        'section': 'section',
        'stream': 'stream',
        'mobile number': 'mobile_number',
        'status': 'status',
    }

    ALLOWED_STREAMS = {'SFM', 'SFW', 'AIDED'}
    TRUE_STATUS_VALUES = {'true', '1', 'yes', 'active', 'enabled'}
    FALSE_STATUS_VALUES = {'false', '0', 'no', 'inactive', 'disabled'}

    def get_queryset(self):
        queryset = (
            Student.objects
            .select_related('user_id', 'academic_department_id')
            .all()
        )

        include_inactive = self.request.query_params.get('include_inactive', '').lower() in {'true', '1', 'yes'}
        status_filter = self.request.query_params.get('status', '').strip()

        if status_filter:
            try:
                parsed_status = self._parse_status_value(status_filter)
            except ValueError:
                queryset = queryset.none()
            else:
                if parsed_status is not None:
                    queryset = queryset.filter(status=parsed_status)
        elif not include_inactive:
            queryset = queryset.filter(status=True)

        exact_filters = {
            'register_number__iexact': self.request.query_params.get('register_number', '').strip(),
            'year_of_admission__iexact': self.request.query_params.get('year_of_admission', '').strip(),
            'section__iexact': self.request.query_params.get('section', '').strip(),
            'stream__iexact': self.request.query_params.get('stream', '').strip(),
        }
        for lookup, value in exact_filters.items():
            if value:
                queryset = queryset.filter(**{lookup: value})

        department_id = self.request.query_params.get('academic_department_id', '').strip()
        if department_id:
            queryset = queryset.filter(academic_department_id_id=department_id)

        return queryset

    @staticmethod
    def _normalize_header(value):
        text = str(value).strip().lower().replace('_', ' ')
        text = ''.join(ch if ch.isalnum() or ch.isspace() else ' ' for ch in text)
        return ' '.join(text.split())

    @staticmethod
    def _parse_optional_date(value):
        if value in (None, ''):
            return None

        if isinstance(value, dt.datetime):
            return value.date()
        if isinstance(value, dt.date):
            return value

        text = str(value).strip()
        if not text:
            return None

        for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y'):
            try:
                return dt.datetime.strptime(text, fmt).date()
            except ValueError:
                continue

        raise ValueError('DOB must be a valid date in DD-MM-YYYY format.')

    @staticmethod
    def _load_rows_from_file(file):
        name = (getattr(file, 'name', '') or '').lower()
        if name.endswith('.csv'):
            file.seek(0)
            text_stream = io.TextIOWrapper(file, encoding='utf-8-sig', newline='')
            reader = csv.reader(text_stream)
            rows = list(reader)
            text_stream.detach()
            return rows

        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        return list(ws.iter_rows(values_only=True))

    @classmethod
    def _extract_header_map(cls, header_row):
        col_map = {}
        if not header_row:
            return col_map, set(cls.HEADER_ALIASES.values())

        for idx, cell in enumerate(header_row):
            if cell is None:
                continue
            normalized = cls._normalize_header(cell)
            field_name = cls.HEADER_ALIASES.get(normalized)
            if field_name and field_name not in col_map:
                col_map[field_name] = idx

        missing = {'register_number', 'name', 'year_of_admission', 'stream', 'mobile_number'}
        missing -= set(col_map.keys())
        return col_map, missing

    @staticmethod
    def _normalize_stream(value):
        if value in (None, ''):
            return None

        normalized = str(value).strip()
        if not normalized:
            return None

        lookup = normalized.upper()
        if lookup == 'AIDED':
            return 'AIDED'
        if lookup in {'SFM', 'SFW'}:
            return lookup

        raise ValueError('Stream must be one of: SFM, SFW, Aided.')

    @staticmethod
    def _parse_status_value(value):
        if value in (None, ''):
            return True

        normalized = str(value).strip().lower()
        if normalized in StudentViewSet.TRUE_STATUS_VALUES:
            return True
        if normalized in StudentViewSet.FALSE_STATUS_VALUES:
            return False

        raise ValueError('Status must be true/false, active/inactive, or enabled/disabled.')

    @staticmethod
    def _extract_department_code(register_number):
        value = str(register_number or '').strip()
        if len(value) < 5:
            return None
        return value[2:5].strip().upper() or None

    def _get_department_from_register_number(self, register_number):
        department_code = self._extract_department_code(register_number)
        if not department_code:
            return None

        return AcademicDepartment.objects.filter(code__iexact=department_code).first()

    def perform_create(self, serializer):
        instance = serializer.save()
        AuditLog.log(
            request=self.request,
            action='CREATE',
            obj=instance,
            changes=self.get_serializer(instance).data,
        )

    def perform_update(self, serializer):
        instance = self.get_object()
        old_data = self.get_serializer(instance).data
        updated_instance = serializer.save()
        new_data = self.get_serializer(updated_instance).data

        changes = {}
        for key, new_value in new_data.items():
            old_value = old_data.get(key)
            if old_value != new_value:
                changes[key] = {'old': old_value, 'new': new_value}

        if changes:
            AuditLog.log(
                request=self.request,
                action='UPDATE',
                obj=updated_instance,
                changes=changes,
            )

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        snapshot = self.get_serializer(instance).data
        object_id = instance.pk

        instance.status = False
        instance.save(update_fields=['status', 'updated_at'])

        if instance.user_id:
            instance.user_id.is_active = False
            instance.user_id.save(update_fields=['is_active', 'updated_at'])

        AuditLog.log(
            request=request,
            action='DEACTIVATE',
            obj=instance,
            object_id=object_id,
            changes=snapshot,
        )
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['get'], url_path='export')
    def export_excel(self, request):
        qs = self.get_queryset().select_related('user_id', 'academic_department_id')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Students'
        ws.append(self.EXCEL_HEADERS)

        for student in qs:
            ws.append([
                student.register_number,
                student.name,
                student.user_id.email if student.user_id else '',
                student.academic_department_id_id or '',
                student.year_of_admission,
                student.dob.strftime('%d-%m-%Y') if student.dob else '',
                student.section or '',
                'Aided' if str(student.stream).upper() == 'AIDED' else student.stream,
                student.mobile_number,
                bool(student.status),
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=students.xlsx'
        wb.save(response)

        AuditLog.log(
            request=request,
            action='EXPORT',
            obj=Student,
            object_id='EXPORT_STUDENTS',
            object_repr='Exported Students',
            changes={'exported_count': qs.count()},
        )

        return response

    def _import_students_from_excel(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            rows = self._load_rows_from_file(file)
        except Exception:
            return Response({'error': 'Invalid Excel file format'}, status=status.HTTP_400_BAD_REQUEST)

        if len(rows) < 2:
            return Response({'error': 'File is empty or contains only headers'}, status=status.HTTP_400_BAD_REQUEST)

        col_map, missing = self._extract_header_map(rows[0])
        if missing:
            missing_readable = ', '.join(sorted(field.replace('_', ' ').title() for field in missing))
            return Response(
                {
                    'error': (
                        f"Missing required header(s): {missing_readable}. "
                        f"Required headers are: {', '.join(self.EXCEL_HEADERS)}"
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        role = Role.objects.filter(name='STUDENT').first()
        if not role:
            return Response({'error': 'STUDENT role does not exist.'}, status=status.HTTP_400_BAD_REQUEST)

        errors = []
        valid_rows = []
        uploaded_register_numbers = []
        skipped_register_numbers = []
        seen_register_numbers = set()
        seen_emails = set()

        existing_register_numbers = {
            str(value).strip().lower()
            for value in Student.objects.values_list('register_number', flat=True)
        }
        existing_register_numbers_report = set()
        existing_usernames = {
            str(value).strip().lower()
            for value in User.objects.values_list('username', flat=True)
        }
        existing_emails = {
            str(value).strip().lower()
            for value in User.objects.exclude(email__isnull=True).exclude(email='').values_list('email', flat=True)
        }

        for row_idx, row in enumerate(rows[1:], start=2):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue

            def get_val(field_name):
                idx = col_map.get(field_name)
                if idx is None or idx >= len(row) or row[idx] is None:
                    return ''
                value = row[idx]
                if isinstance(value, str):
                    return value.strip()
                return value

            register_number = str(get_val('register_number')).strip()
            name = str(get_val('name')).strip()
            email = str(get_val('email')).strip()
            year_of_admission = str(get_val('year_of_admission')).strip()
            dob_raw = get_val('dob')
            section = str(get_val('section')).strip()
            stream_raw = get_val('stream')
            mobile_number = str(get_val('mobile_number')).strip()
            status_raw = get_val('status')

            row_errors = []

            if not register_number:
                row_errors.append('Register Number is required.')
            elif len(register_number) > 30:
                row_errors.append('Register Number cannot exceed 30 characters.')
            else:
                register_key = register_number.lower()
                if register_key in existing_register_numbers or register_key in existing_usernames:
                    if register_key not in existing_register_numbers_report:
                        skipped_register_numbers.append(register_number)
                        existing_register_numbers_report.add(register_key)
                    continue
                if register_key in seen_register_numbers:
                    if register_key not in existing_register_numbers_report:
                        skipped_register_numbers.append(register_number)
                        existing_register_numbers_report.add(register_key)
                    continue
                seen_register_numbers.add(register_key)

            if not name:
                row_errors.append('Name is required.')
            elif len(name) > 100:
                row_errors.append('Name cannot exceed 100 characters.')

            if email:
                try:
                    validate_email(email)
                except DjangoValidationError:
                    row_errors.append(f"Email '{email}' is not valid.")
                else:
                    email_key = email.lower()
                    if email_key in existing_emails:
                        row_errors.append(f"Email '{email}' already exists.")
                    elif email_key in seen_emails:
                        row_errors.append(f"Duplicate Email '{email}' found within the uploaded Excel file.")
                    else:
                        seen_emails.add(email_key)

            department = self._get_department_from_register_number(register_number)
            if not department:
                row_errors.append('Department not found.')

            if not year_of_admission:
                row_errors.append('Year of Admission is required.')
            elif len(year_of_admission) > 9:
                row_errors.append('Year of Admission cannot exceed 9 characters.')

            dob = None
            try:
                dob = self._parse_optional_date(dob_raw)
            except ValueError as exc:
                row_errors.append(str(exc))
            if not dob:
                row_errors.append('DOB is required.')

            if section and len(section) > 10:
                row_errors.append('Section cannot exceed 10 characters.')

            try:
                stream = self._normalize_stream(stream_raw)
            except ValueError as exc:
                row_errors.append(str(exc))
                stream = None

            if not mobile_number:
                row_errors.append('Mobile Number is required.')
            elif len(mobile_number) > 15:
                row_errors.append('Mobile Number cannot exceed 15 characters.')

            try:
                status_value = self._parse_status_value(status_raw)
            except ValueError as exc:
                row_errors.append(str(exc))
                status_value = True

            if row_errors:
                errors.append({'row': row_idx, 'errors': row_errors})
                continue

            valid_rows.append(
                {
                    'register_number': register_number,
                    'name': name,
                    'email': email or None,
                    'department': department,
                    'year_of_admission': year_of_admission,
                    'dob': dob,
                    'section': section or None,
                    'stream': stream,
                    'mobile_number': mobile_number,
                    'status': status_value,
                }
            )

        if errors:
            return Response(
                {
                    'error': 'Validation failed for some rows',
                    'details': errors,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not valid_rows:
            if skipped_register_numbers:
                return Response(
                    {
                        'message': 'already existing data',
                        'already_existing_register_numbers': skipped_register_numbers,
                        'uploaded_register_numbers': [],
                    },
                    status=status.HTTP_200_OK,
                )
            return Response(
                {'message': 'No valid student records found.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created_count = 0
        with transaction.atomic():
            for row in valid_rows:
                user = User.objects.create(
                    username=row['register_number'],
                    email=row['email'],
                    password_hash=make_password(row['dob'].isoformat()),
                    role_id=role,
                    is_active=bool(row['status']),
                )

                Student.objects.create(
                    register_number=row['register_number'],
                    name=row['name'],
                    dob=row['dob'],
                    user_id=user,
                    academic_department_id=row['department'],
                    year_of_admission=row['year_of_admission'],
                    section=row['section'],
                    stream=row['stream'],
                    mobile_number=row['mobile_number'],
                    status=row['status'],
                )
                created_count += 1
                uploaded_register_numbers.append(row['register_number'])

            AuditLog.log(
                request=request,
                action='IMPORT',
                obj=Student,
                object_id='BULK_IMPORT',
                object_repr='Imported Students',
                changes={'imported_count': created_count},
            )

        if skipped_register_numbers:
            message = (
                f"Uploaded the new student records: {', '.join(uploaded_register_numbers)}. "
                f"Already existing register numbers: {', '.join(skipped_register_numbers)}."
            )
        else:
            message = f"Uploaded the new student records: {', '.join(uploaded_register_numbers)}."

        return Response(
            {
                'message': message,
                'uploaded_register_numbers': uploaded_register_numbers,
                'already_existing_register_numbers': skipped_register_numbers,
                'imported_count': created_count,
            },
            status=status.HTTP_201_CREATED,
        )

    @action(detail=False, methods=['post'], url_path='import', parser_classes=[MultiPartParser, FormParser])
    def import_excel(self, request):
        return self._import_students_from_excel(request)

    @action(detail=False, methods=['post'], url_path='bulk-import', parser_classes=[MultiPartParser, FormParser])
    def bulk_import_excel(self, request):
        return self._import_students_from_excel(request)
